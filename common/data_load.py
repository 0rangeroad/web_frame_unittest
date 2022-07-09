import os
import openpyxl
import yaml

dir_path = os.path.dirname(__file__).split('common')[0]


def read_extract_yaml(key):
    with open(dir_path + "extract.yaml", 'r', encoding='utf-8') as f:
        value = yaml.load(stream=f, Loader=yaml.FullLoader)
        return value[key]


# 'a',追加数据，但数据会累计
def write_extract_yaml(data):
    with open(dir_path + "extract.yaml", 'a', encoding='utf-8') as f:
        yaml.dump(data=data, stream=f, allow_unicode=True)


def clear_extract_yaml():
    with open(dir_path + "extract.yaml", 'w', encoding='utf-8') as f:
        f.truncate()


def read_testcase_yaml(yaml_name):
    print(os.getcwd() + "/data/")
    with open(dir_path + "data/" + yaml_name, 'r', encoding='utf-8') as f:
        value = yaml.load(stream=f, Loader=yaml.FullLoader)
        return value


def read_excel_list():
    wb = openpyxl.load_workbook(dir_path + 'data/test_data.xlsx')
    sheet = wb['login']
    print('最大行', sheet.max_row)
    print('最大列', sheet.max_column)
    all_list = []
    for rows in range(2, sheet.max_row + 1):
        temp_list = []
        for cols in range(1, sheet.max_column + 1):
            if sheet.cell(rows, cols).value is not None:
                temp_list.append(sheet.cell(rows, cols).value)
        all_list.append(temp_list)
    return all_list


def read_excel_dic():
    wb = openpyxl.load_workbook(dir_path + 'data/test_data.xlsx')
    sheet = wb['1']
    res = list(sheet.rows)
    # 获取第一行表头
    title = [i.value for i in res[0]]
    # 遍历第一行之外的其他行
    cases = []
    for item in res[1:]:
        data = [i.value for i in item]
        dic = dict(zip(title, data))
        cases.append(dic)
    return cases


def write_excel(row, column, value):
    wb = openpyxl.load_workbook(dir_path + 'data/test_data.xlsx')
    sheet = wb['1']
    sheet.cell(row=row, column=column, value=value)
    wb.save(dir_path + 'data/test_data.xlsx')
    wb.close()
